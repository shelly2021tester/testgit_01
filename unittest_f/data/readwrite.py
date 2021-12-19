'''
@File:readwrite.py
@DateTime:2021/12/17 15:26
@Author:shelly
@Desc:
'''
import openpyxl

class Read_Write:
    def __init__(self,file,sheet):
        self.file=file
        self.wb = openpyxl.load_workbook(self.file)
        table = self.wb[sheet]
        self.table = self.wb.active
        self.nrows = table.max_row
        self.ncols = table.max_column

    def read(self):
        list2=[]
        for i in range(2,self.nrows+1):
            list1 = []
            for j in range(1,self.ncols+1):
                content=self.table .cell(i,j).value
                list1.append(content)
            list2.append(list1)
        self.wb.close()
        return list2

    def write(self,*arg):
        for i in range(self.nrows+1,self.nrows+2):
            for j in range(1,len(arg)+1):
                self.table.cell(i,j).value=arg[j-1]
        self.wb.save(self.file)
        self.wb.close()

if __name__=='__main__':
    file=r"C:\Users\zxl\Desktop\test.xlsx"
    sheet="login"
    doc1=Read_Write(file,sheet)
    # print(doc1.read())
    doc1.write("test","aa","das","asd")